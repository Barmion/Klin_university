import io

import pytest
from faker import Faker
from openpyxl import Workbook
from rest_framework.test import APIClient
from tests.factories import PositionFactory, UserFactory, WorkerFactory

fake = Faker()


@pytest.fixture
def admin_user(django_user_model):
    return django_user_model.objects.create_user(
        username='admin',
        password='admin',
        is_staff=True,
    )


@pytest.fixture
def api_client():
    client = APIClient()
    return client


@pytest.fixture
def admin_api_client(admin_user):
    client = APIClient()
    client.force_authenticate(user=admin_user)
    return client


@pytest.fixture
def position():
    return PositionFactory()


@pytest.fixture
def user():
    return UserFactory()


@pytest.fixture
def worker(user, position):
    return WorkerFactory(created_by=user, position=position)


@pytest.fixture
def workers(user, position):
    def inner(n):
        return WorkerFactory.create_batch(n, created_by=user, position=position)
    return inner


@pytest.fixture
def data(position):
    return {
        'first_name': fake.first_name(),
        'last_name': fake.last_name(),
        'email': fake.email(),
        'position': position.name,
    }


@pytest.fixture
def excel_file(data):
    wb = Workbook()
    ws = wb.active
    ws.append(list(data.keys()))
    ws.append(list(data.values()))
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    excel_file.name = "test_workers.xlsx"
    return excel_file


@pytest.fixture
def bad_excel_file(data):
    wb = Workbook()
    ws = wb.active
    ws.append(list(data.keys()))
    ws.append(['Name', 'Lastname', 'bademail', data['position']])
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    excel_file.name = "test_workers.xlsx"
    return excel_file


@pytest.fixture
def mixed_excel_file(data):
    wb = Workbook()
    ws = wb.active
    ws.append(list(data.keys()))
    ws.append(list(data.values()))
    ws.append(['Name', 'Lastname', 'bademail', data['position']])
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    excel_file.name = "test_workers.xlsx"
    return excel_file