from rest_framework.exceptions import ValidationError
from openpyxl import load_workbook


def validate_uploaded_file(request):
    if 'file' not in request.FILES:
        raise ValidationError({
            'error': 'Файл не предоставлен',
            'message': 'Передайте файл в поле "file"'
        })

    file = request.FILES['file']

    if not file.name.endswith(('.xlsx', '.xls')):
        raise ValidationError({
            'error': 'Неверный формат файла',
            'message': 'Поддерживаются только Excel файлы (.xlsx, .xls)'
        })

    return file


def parse_excel(file):
    workbook = load_workbook(filename=file, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    return [
        dict(zip(headers, row))
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]
